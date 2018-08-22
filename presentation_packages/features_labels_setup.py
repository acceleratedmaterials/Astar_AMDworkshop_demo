import numpy as np
import numpy.random as rng
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import StratifiedKFold, KFold, LeaveOneOut, train_test_split
from keras.utils import to_categorical
import pickle
import os
import pathlib

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.expand_frame_repr', False)


def print_array_to_excel(array, first_cell, ws, axis=2):
    '''
    Print an np array to excel using openpyxl
    :param array: np array
    :param first_cell: first cell to start dumping values in
    :param ws: worksheet reference. From openpyxl, ws=wb[sheetname]
    :param axis: to determine if the array is a col vector (0), row vector (1), or 2d matrix (2)
    '''
    if isinstance(array, (list,)):
        array = np.array(array)
    shape = array.shape
    if axis == 0:
        # Treat array as col vector and print along the rows
        array.flatten()  # Flatten in case the input array is a nx1 ndarry which acts weird
        for i in range(shape[0]):
            j = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i]
    elif axis == 1:
        # Treat array as row vector and print along the columns
        array.flatten()  # Flatten in case the input array is a 1xn ndarry which acts weird
        for j in range(shape[0]):
            i = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[j]
    elif axis == 2:
        # If axis==2, means it is a 2d array
        for i in range(shape[0]):
            for j in range(shape[1]):
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]


def read_reaction_data(loader_excel_file='./excel/data_loader.xlsx', mode='c', save_mode=False, print_mode=False):
    '''
    Used to read reaction data from excel file containing reaction data. Must split the workbook into 4 sheets
    df : sheet containing the headers and values for all reaction data
    features_c: sheet containing continuous features values. Each row is one example with each col being diff. features
    features_d: sheet containing categorical features. As of now, the code can only handle 1 categorical feature. :(
    :param loader_excel_file: excel file name
    :param print_mode: Print results
    :return: fl class object
    '''
    try:
        df = pd.read_excel(loader_excel_file, sheet_name='df')
    except FileNotFoundError:
        try:
            # Check for xlsx file with the same name but with file extension missing
            if loader_excel_file[-5:] != '.xlsx':
                # Must create a temp name because there is another check after this
                loader_excel_file1 = loader_excel_file + '.xlsx'
                df = pd.read_excel(loader_excel_file1, sheet_name='df')
            else:
                # Means that even with .xlsx at the back, file is not found, so raise an error.
                raise FileNotFoundError
        except FileNotFoundError:
            # Check for xlmx file instead
            if loader_excel_file[-5:] != '.xlsm':
                loader_excel_file = loader_excel_file + '.xlsm'
            df = pd.read_excel(loader_excel_file, sheet_name='df')
        else:
            # If first check succeeds, rename excel file name with the temp file name
            loader_excel_file = loader_excel_file1
    features_c = pd.read_excel(loader_excel_file, sheet_name='features_c').values
    # There should be one column in the excel sheet for labels only!
    if mode == 'c':
        labels = pd.read_excel(loader_excel_file, sheet_name='labels').values.flatten()
    elif mode == 'r':
        labels = pd.read_excel(loader_excel_file, sheet_name='labels').values
    else:
        raise TypeError('mode input should be either c, r, or p')
    fl = Features_labels(features_c, labels, mode=mode, save_mode=save_mode)
    if print_mode:
        print(df)
    return fl


class Features_labels:
    def __init__(self, features_c, labels, mode='c', scaler=None, num_classes=None, save_name='fl', save_mode=False):
        """
        Creates fl class with a lot useful attributes
        :param features_c: Continuous features. Np array, no. of examples x continous features
        :param labels: Labels as np array, no. of examples x 1
        :param mode: c for classification mode, r for regression mode
        :param scaler: Scaler to transform features c. If given, use given MinMax scaler from sklearn,
        else create scaler based on given features c.
        :param save_name: If save mode is on, will be the name of the saved fl class obj
        :param save_mode: To save or not.
        """

        def features_to_listedtuple(features, targets):
            '''
            Converts features from np array to list of sorted tuples based on labels
            :param features: features to be sorted
            :param targets: labels
            :return: list of sorted tuples. eg: [(0, features nxm), (1, features nxm)]
            '''
            dic = {}
            for feature, target in zip(np.ndarray.tolist(features), targets):
                if target in dic:  # Create new class tuple in the dic
                    dic[target].append(feature)
                else:  # If class already exists, append new features into that class
                    dic[target] = [feature]
            for target in dic:  # Convert list back to ndarray
                dic[target] = np.array(dic[target])
            # Convert from dictionary to list of tuple
            return sorted(dic.items())

        self.mode = mode
        if mode == 'c':  # Means classification mode
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Without _a at the back means it is the listed tuple data type.
            self.features_c = features_to_listedtuple(features_c, labels)
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            self.features_c_norm = features_to_listedtuple(self.features_c_norm_a, labels)
            # Setting up labels
            self.labels = labels
            if num_classes is None:
                _, count = np.unique(labels, return_counts=True)
                self.n_classes = len(count)
            else:
                self.n_classes = num_classes
            self.labels_hot = to_categorical(labels, num_classes=self.n_classes)
            # List containing number of examples per class
            self.count_per_class = [category[1].shape[0] for category in self.features_c]
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]
        elif mode == 'r':  # Means regression mode
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            # Setting up labels
            self.labels = labels
            assert num_classes is None, 'If regression mode is selected, num_classes should be None'
            self.n_classes=None
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]
            if len(labels.shape) == 2:
                self.labels_dim = labels.shape[1]
            else:
                self.labels_dim = 1
        elif mode == 'p':  # Means prediction mode
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
        else:
            raise TypeError('mode given should be either c, r, or p')

        # Saving
        if save_mode:
            file_path = open('./save/features_labels/' + save_name + '.obj', 'wb')
            pickle.dump(self, file_path)

    def generate_random_examples(self, numel):
        gen_features_c_norm_a = rng.random_sample((numel, self.features_c_dim))
        gen_features_c_a = self.scaler.inverse_transform(gen_features_c_norm_a)
        gen_fl=Features_labels(gen_features_c_a, labels=None, mode='p', scaler=self.scaler, num_classes=self.n_classes,
                               save_mode=False)
        return gen_fl

    def generate_random_subset(self, subset_split, save_fl=False, save_to_excel=False,
                               loader_excel_file='./excel/data_loader.xlsx'):
        '''
        Split main data set of examples into subsets containing the desired number of classes. For example,
        subset_split = [3,3,3] means make a subset with 3 class 0s examples, 3 class 1s, 3 class 2s
        :param subset_split: List with number of elements equal to total number of class
        :param loader_excel_file: Name of data loader excel file to be open and to save stuff to
        :return:
        '''
        # Setting up subset examples
        for category in range(self.n_classes):
            try:
                # Subset
                n_examples = self.features_c_norm[category][1].shape[0]  # Get no. of examples for current class
                idx = rng.choice(n_examples, size=(subset_split[category],), replace=False)
                features_c_a = self.features_c[category][1][idx, :]  # Choose number of examples for current class
                labels = np.repeat(category, subset_split[category])  # Set the labels for the selected examples
                # Inverse subset. Deleting the examples chosen for the subset
                i_features_c_a = np.delete(self.features_c[category][1], idx, axis=0)
                i_labels = np.repeat(category, n_examples - subset_split[category])
                if category == 0:  # Means first loop and the array is not formed yet
                    features_c_a_store = features_c_a
                    labels_store = labels
                    i_features_c_a_store = i_features_c_a
                    i_labels_store = i_labels
                else:
                    features_c_a_store = np.concatenate((features_c_a_store, features_c_a), axis=0)
                    labels_store = np.concatenate((labels_store, labels))
                    i_features_c_a_store = np.concatenate((i_features_c_a_store, i_features_c_a), axis=0)
                    i_labels_store = np.concatenate((i_labels_store, i_labels))
            except:
                if len(subset_split) != self.n_classes:
                    raise Exception('Subset split does not have same number of elements as the total number of class!'
                                    'Make sure that they are the same.')
                continue
        ss_fl = Features_labels(features_c_a_store, labels_store, save_name='ss_fl', save_mode=save_fl)
        # Setting up inverse subset examples. Note: Must use ss_scaler to ensure that when doing the prediction, the
        # inverse subset is scaled correctly and the same as when the model is trained with the subset examples
        ss_scaler = ss_fl.scaler
        i_ss_fl = Features_labels(i_features_c_a_store, i_labels_store, scaler=ss_scaler,
                                  save_name='i_ss_fl', save_mode=save_fl)
        # Excel writing part
        if save_to_excel:
            wb = load_workbook(loader_excel_file)
            # Setting up subset features and labels sheet
            sheet_name_store = ['ss_features_c', 'ss_labels', 'i_ss_features_c', 'i_ss_labels']
            ss_store = [features_c_a_store, labels_store, i_features_c_a_store, i_labels_store]
            axis_store = [2, 0, 2, 0]  # Because feature_c is 2D while labels is col vector, so order of axis is 2,0,2,0
            for cnt, sheet_name in enumerate(sheet_name_store):
                if sheet_name in wb.sheetnames:
                    # If temp sheet exists, remove it to create a new one. Else skip this.
                    idx = wb.sheetnames.index(sheet_name)  # index of temp sheet
                    wb.remove(wb.worksheets[idx])  # remove temp
                    wb.create_sheet(sheet_name, idx)  # create an empty sheet using old index
                else:
                    wb.create_sheet(sheet_name)  # Create the new sheet
                # Print array to the correct worksheet
                print_array_to_excel(ss_store[cnt], (2, 1), wb[sheet_name_store[cnt]], axis=axis_store[cnt])
            wb.save(loader_excel_file)
            wb.close()
        return ss_fl, i_ss_fl

    def create_skf(self, k_folds):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for stratified k fold cross validation
        :param k_folds: Number of folds
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = StratifiedKFold(n_splits=k_folds, shuffle=True)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False),
                 Features_labels(xval, yval, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False))
            )
        return fl_store

    def create_kf(self, k_folds):
        '''
        Almost the same as skf except can work for regression labels and folds are not stratified.
        Create list of tuples containing (fl_train,fl_val) fl objects for k fold cross validation
        :param k_folds: Number of folds
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=False)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False),
                 Features_labels(xval, yval, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False))
            )
        return fl_store

    def create_loocv(self):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for leave one out cross validation
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        loocv = LeaveOneOut()
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(loocv.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False),
                 Features_labels(xval, yval, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False))
            )
        return fl_store

    def create_train_test_split(self, test_size=0.2):
        x_train, x_test, y_train, y_test = train_test_split(self.features_c_a, self.labels, test_size=test_size)
        train_fl = Features_labels(x_train, y_train, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                   save_mode=False)
        test_fl = Features_labels(x_test, y_test, mode=self.mode, scaler=self.scaler, num_classes=self.n_classes,
                                  save_mode=False)
        return train_fl, test_fl

