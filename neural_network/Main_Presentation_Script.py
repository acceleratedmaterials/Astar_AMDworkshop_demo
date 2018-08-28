import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error

from presentation_packages.features_labels_setup import read_reaction_data
from presentation_packages.DNN_setup import create_hparams, DNN
from presentation_packages.ODE import denbigh_rxn2

# Read excel data into fl class
fl = read_reaction_data(loader_excel_file='./excel/data_loader/denbigh_data_loader.xlsx', mode='r')
train_fl, test_fl = fl.create_train_test_split(test_size=0.2)

# Create DNN class
hparams = create_hparams(hidden_layers=[50, 50], epochs=100, batch_size=64, activation='relu')
dnn_class = DNN(hparams=hparams, fl=fl)
dnn_class.train_model(fl=train_fl)
_, mse = dnn_class.eval(eval_fl=test_fl, mode='r')
print('mse of test set is {}'.format(mse))

# Generate random examples
gen_fl = fl.generate_random_examples(numel=1000)
predictions = dnn_class.eval(eval_fl=gen_fl, mode='p')

# Actual output concentrations using ODE to calculate
actual_conc = denbigh_rxn2(gen_fl.features_c_a)
mse_diff=mean_squared_error(actual_conc,predictions)
print('mse between actual concentration and predicted concentration is {}'.format(mse_diff))

# Printing generated examples into an excel file
# Making a df containing all the information
headers = ['f' + str(idx + 1) for idx in range(fl.features_c_count)] + ['P' + str(idx + 1) for idx in
                                                                        range(fl.labels_dim)] \
          + ['A' + str(idx + 1) for idx in range(fl.labels_dim)]
df = np.concatenate((gen_fl.features_c_a, predictions, actual_conc), axis=1)
df = pd.DataFrame(data=df, columns=headers)

# Opening excel workbook
gen_excel_name = './excel/gen/gen.xlsx'
gen_sheet_name = 'gen'
wb = load_workbook(gen_excel_name)

# Creating new worksheet
wb.create_sheet('gen')
sheet_name = wb.sheetnames[-1]  # Taking the ws name from the back to ensure the newly created sheet is selected

# Print df to excel
pd_writer = pd.ExcelWriter(gen_excel_name, engine='openpyxl')
pd_writer.book = wb
pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
df.to_excel(pd_writer, sheet_name)
pd_writer.save()
pd_writer.close()
wb.close()
